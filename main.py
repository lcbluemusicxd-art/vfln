import requests
from requests.auth import HTTPDigestAuth
import urllib3
import pandas as pd
import os
import time
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============= PARÁMETROS DE TOLERANCIA Y HORAS EXTRA =============
TOLERANCIA_ENTRADA_TARDE = 10  # minutos de tolerancia después de hora de entrada
TOLERANCIA_SALIDA_ANTES = 5    # minutos antes de hora de salida
TOLERANCIA_SALIDA_DESPUES = 40 # minutos después de hora de salida
HORA_LIMITE_EXTRA_LJ = "19:10" # Hora común de horas extra Lunes-Jueves (puede llegar hasta 19:20)
TOLERANCIA_HORA_EXTRA = 10     # minutos de tolerancia para hora límite extra
MAX_HORAS_EXTRA = 2            # máximo 2 horas extra por día
HORAS_EXTRA_VIERNES = False    # False = no hay horas extra los viernes

# Configuración SÁBADOS
SABADO_HORA_LIMITE_SALIDA = "17:10"  # Salida máxima sábados (incluso con horas extra)
SABADO_TOLERANCIA_SALIDA = 10         # minutos de tolerancia para salida
SABADO_ENTRADA_MINIMA_REAL = "06:10"  # Si entra antes de esta hora, usar hora real

# Configuración SERVICIOS GENERALES
SERVICIOS_TOLERANCIA_ENTRADA = 60     # 1 hora de tolerancia entrada
SERVICIOS_HORA_SALIDA_MIN = "14:00"   # Puede salir desde las 2 PM

# Configuración DESAYUNO Y COMIDA (para cálculo de nómina)
DESAYUNO_HORA = "09:00"               # Hora del desayuno
DESAYUNO_MINUTOS_DESCUENTO = 10       # Minutos a descontar por desayuno
COMIDA_HORA_INICIO = "13:00"          # Hora inicio de comida
COMIDA_DURACION_MINUTOS = 30          # Duración de comida (minutos a descontar)
COMIDA_TOLERANCIA_SALIDA = 15         # Tolerancia en minutos para salir antes de comida sin descuento

IP_RELOJ = "192.168.1.97"
USER = "admin"
PASS = "VaferluJL17*"
CARPETA_FOTOS = "fotos_reloj"
MAX_WORKERS = 10

if not os.path.exists(CARPETA_FOTOS):
    os.makedirs(CARPETA_FOTOS)


def obtener_tolerancias_para_empleado(id_empleado):
    """Obtiene las tolerancias a usar para un empleado (personalizadas o globales)"""
    from jornada_laboral import obtener_tolerancias_empleado
    
    tol_personal = obtener_tolerancias_empleado(id_empleado)
    
    if tol_personal:
        # Tiene tolerancias personalizadas
        return {
            'tolerancia_entrada': tol_personal[1],
            'tolerancia_salida_antes': tol_personal[2],
            'tolerancia_salida_despues': tol_personal[3],
            'hora_limite_extra_lj': tol_personal[4],
            'tolerancia_hora_extra': tol_personal[5],
            'horas_extra_maximo': tol_personal[6],
            'viernes_permite_extra': bool(tol_personal[7]),
            'sabado_hora_limite': tol_personal[8],
            'sabado_tolerancia': tol_personal[9],
            'sabado_entrada_minima': tol_personal[10],
            'servicios_tolerancia_entrada': tol_personal[11],
            'servicios_hora_salida_min': tol_personal[12],
            'comida_tolerancia_salida': tol_personal[14] if len(tol_personal) > 14 else COMIDA_TOLERANCIA_SALIDA
        }
    else:
        # Usar tolerancias globales
        return {
            'tolerancia_entrada': TOLERANCIA_ENTRADA_TARDE,
            'tolerancia_salida_antes': TOLERANCIA_SALIDA_ANTES,
            'tolerancia_salida_despues': TOLERANCIA_SALIDA_DESPUES,
            'hora_limite_extra_lj': HORA_LIMITE_EXTRA_LJ,
            'tolerancia_hora_extra': TOLERANCIA_HORA_EXTRA,
            'horas_extra_maximo': MAX_HORAS_EXTRA,
            'viernes_permite_extra': HORAS_EXTRA_VIERNES,
            'sabado_hora_limite': SABADO_HORA_LIMITE_SALIDA,
            'sabado_tolerancia': SABADO_TOLERANCIA_SALIDA,
            'sabado_entrada_minima': SABADO_ENTRADA_MINIMA_REAL,
            'servicios_tolerancia_entrada': SERVICIOS_TOLERANCIA_ENTRADA,
            'servicios_hora_salida_min': SERVICIOS_HORA_SALIDA_MIN,
            'comida_tolerancia_salida': COMIDA_TOLERANCIA_SALIDA
        }


def obtener_nombres_empleados(auth):
    """Obtiene el mapa de nombres para el Excel con paginacion."""
    url = f"https://{IP_RELOJ}/ISAPI/AccessControl/UserInfo/Search?format=json"
    nombres_map = {}
    posicion = 0
    bloque = 100

    while True:
        payload = {
            "UserInfoSearchCond": {
                "searchID": "user_sync_" + datetime.now().strftime("%H%M%S"),
                "searchResultPosition": posicion,
                "maxResults": bloque
            }
        }
        try:
            res = requests.post(url, json=payload, auth=auth, verify=False, timeout=15)
            if res.status_code == 200:
                data_response = res.json().get('UserInfoSearch', {})
                data = data_response.get('UserInfo', [])

                if not data:
                    logger.info(f"No mas empleados. Total descargados: {len(nombres_map)}")
                    break

                for u in data:
                    id_str = str(u.get('employeeNoString') or u.get('employeeNo') or '')
                    nombre = u.get('name', 'Sin Nombre')

                    if id_str:
                        nombres_map[id_str] = nombre

                logger.info(f"Pagina descargada: {len(data)} empleados. Total acumulado: {len(nombres_map)}")

                total_matches = data_response.get('totalMatches', 0)
                if posicion + len(data) >= total_matches:
                    logger.info(f"Se cargaron todos los {len(nombres_map)} empleados de la base de datos")
                    break

                posicion += len(data)

            else:
                logger.error(f"Error al obtener nombres: Status {res.status_code}")
                break

        except requests.exceptions.Timeout:
            logger.error("Timeout al conectar con el reloj para obtener nombres")
            break
        except requests.exceptions.ConnectionError as e:
            logger.error(f"Error de conexion al reloj: {e}")
            break
        except ValueError as e:
            logger.error(f"Error al parsear JSON de nombres: {e}")
            break
        except Exception as e:
            logger.error(f"Error inesperado obteniendo nombres: {type(e).__name__} - {e}")
            break

    if '00001010' in nombres_map:
        logger.info(f"ID 00001010 encontrado: {nombres_map['00001010']}")

    return nombres_map


def descargar_foto(url_pic, nombre_archivo, auth):
    """Descarga una foto de forma segura con manejo de errores."""
    ruta_completa = os.path.join(CARPETA_FOTOS, nombre_archivo)

    if os.path.exists(ruta_completa):
        return True, "Ya existe"

    path_limpio = url_pic.split(IP_RELOJ)[-1] if IP_RELOJ in url_pic else url_pic
    if not path_limpio.startswith("/"):
        path_limpio = "/" + path_limpio

    url_final = f"https://{IP_RELOJ}{path_limpio}"

    try:
        res = requests.get(url_final, auth=auth, verify=False, timeout=10)

        if res.status_code == 200:
            try:
                with open(ruta_completa, 'wb') as f:
                    f.write(res.content)
                logger.debug(f"Foto descargada: {nombre_archivo}")
                return True, "Descargada"
            except IOError as e:
                logger.error(f"Error al escribir foto {nombre_archivo}: {e}")
                return False, f"Error I/O: {e}"
        else:
            logger.warning(f"Status {res.status_code} descargando {nombre_archivo}")
            return False, f"HTTP {res.status_code}"

    except requests.exceptions.Timeout:
        logger.warning(f"Timeout descargando {nombre_archivo}")
        return False, "Timeout"
    except requests.exceptions.ConnectionError as e:
        logger.warning(f"Error conexion descargando {nombre_archivo}: {e}")
        return False, "Connection Error"
    except requests.exceptions.RequestException as e:
        logger.error(f"Error de request descargando {nombre_archivo}: {e}")
        return False, f"Request Error: {e}"
    except Exception as e:
        logger.error(f"Error inesperado descargando {nombre_archivo}: {type(e).__name__} - {e}")
        return False, f"Error: {type(e).__name__}"


def descargar_fotos_paralelo(fotos_pendientes, auth):
    """Descarga multiples fotos en paralelo usando ThreadPoolExecutor."""
    logger.info(f"Iniciando descarga paralela de {len(fotos_pendientes)} fotos ({MAX_WORKERS} hilos)...")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(descargar_foto, foto['url'], foto['archivo'], auth): foto
            for foto in fotos_pendientes
        }

        completadas = 0
        for future in as_completed(futures):
            foto = futures[future]
            try:
                exito, mensaje = future.result()
                completadas += 1

                if completadas % 10 == 0:
                    logger.info(f"Progreso: {completadas}/{len(fotos_pendientes)} fotos descargadas")

            except Exception as e:
                logger.error(f"Excepcion en descarga de {foto['archivo']}: {e}")

    logger.info(f"Descarga completada: {completadas}/{len(fotos_pendientes)} fotos procesadas")


def procesar_entrada_salida(lista_asistencia, minutos_salida_minima=30):
    """
    Agrupa los registros por ID y dia, tomando:
    - Primer registro = Hora de Entrada
    - Último registro = Hora de Salida (debe haber pasado al menos minutos_salida_minima desde la entrada)
    - Si hay múltiples registros después de la entrada, valida que tengan >= minutos_salida_minima de diferencia
    - Detecta y registra PERMISOS: un ingreso por PIN = permiso (salida o entrada con permiso)
    - Calcula horas trabajadas considerando segundos, pero mostrando solo HH:MM
    
    Args:
        lista_asistencia: Lista de registros de asistencia
        minutos_salida_minima: Minutos mínimos entre entrada y salida (default 30)
    """
    from jornada_laboral import guardar_permiso
    
    for reg in lista_asistencia:
        try:
            reg['fecha_obj'] = pd.to_datetime(reg['Fecha y Hora']).tz_localize(None)
        except Exception as e:
            logger.warning(f"No se pudo parsear fecha '{reg['Fecha y Hora']}': {e}")
            reg['fecha_obj'] = None

    lista_valida = [r for r in lista_asistencia if r['fecha_obj'] is not None]
    logger.info(f"Se procesaran {len(lista_valida)}/{len(lista_asistencia)} registros con fecha valida")

    datos_agrupados = defaultdict(lambda: {"nombre": "", "dias": {}})

    for reg in lista_valida:
        id_emp = reg['ID']
        nombre = reg['Nombre']
        fecha_obj = reg['fecha_obj']
        archivo_foto = reg['Archivo Foto']
        tipo_acceso = reg.get('Tipo de Acceso', 'Rostro')

        dia_key = fecha_obj.strftime("%Y-%m-%d")
        hora_str = fecha_obj.strftime("%H:%M")

        datos_agrupados[id_emp]['nombre'] = nombre

        # ===== DETECTAR PERMISOS (Ingreso por PIN) =====
        if tipo_acceso.upper() == 'PIN':
            # Un ingreso por PIN = Permiso
            # Determinar si es permiso de salida o entrada
            try:
                if dia_key in datos_agrupados[id_emp]['dias']:
                    # Si ya hay entrada en el día, es permiso de salida
                    tipo_permiso = 'salida'
                else:
                    # Si es el primer ingreso del día, es permiso de entrada (regreso)
                    tipo_permiso = 'entrada'
                
                guardar_permiso(id_emp, dia_key, tipo_permiso, hora_str)
                logger.info(f"Permiso registrado: {id_emp} - {tipo_permiso} a {hora_str}")
            except Exception as e:
                logger.warning(f"Error registrando permiso: {e}")
            
            # No procesar como entrada/salida regular
            continue

        if dia_key not in datos_agrupados[id_emp]['dias']:
            datos_agrupados[id_emp]['dias'][dia_key] = {
                'fecha_obj': fecha_obj,
                'hora_entrada': fecha_obj.strftime("%H:%M"),
                'hora_salida': None,
                'tipo_acceso': tipo_acceso,
                'foto_id': archivo_foto,
                'entrada_datetime': fecha_obj,
                'horas_trabajadas': None,
                'registros_dia': [fecha_obj]  # Guardar todos los registros del día
            }
        else:
            # Agregar este registro a la lista de registros del día
            datos_agrupados[id_emp]['dias'][dia_key]['registros_dia'].append(fecha_obj)
            
            hora_entrada_dt = datos_agrupados[id_emp]['dias'][dia_key]['entrada_datetime']
            
            # Verificar si este registro está lo suficientemente alejado del anterior
            # para considerarlo como una nueva salida válida
            diferencia = fecha_obj - hora_entrada_dt
            minutos_diferencia = diferencia.total_seconds() / 60
            
            # El registro es válido como salida si ha pasado al menos minutos_salida_minima
            if minutos_diferencia >= minutos_salida_minima:
                segundos_totales = int(diferencia.total_seconds())
                horas = segundos_totales // 3600
                minutos = (segundos_totales % 3600) // 60
                segundos_restantes = segundos_totales % 60
                
                if segundos_restantes >= 30:
                    minutos += 1
                    if minutos == 60:
                        horas += 1
                        minutos = 0
                
                datos_agrupados[id_emp]['dias'][dia_key]['hora_salida'] = fecha_obj.strftime("%H:%M")
                datos_agrupados[id_emp]['dias'][dia_key]['horas_trabajadas'] = f"{horas:02d}:{minutos:02d}"

    return datos_agrupados


def generar_excel_final(datos_agrupados):
    """
    Estructura del Excel - Reporte Detallado sin Descuentos
    Fila 1: ID | Nombre | Lunes 5 (fusionado 6 cols) | Martes 6 (fusionado 6 cols) | ... 
    Fila 2:    |        | Entrada|Salida|Horas|Acceso|Foto|Asistencia | Entrada|Salida|Horas|Acceso|Foto|Asistencia | ... |
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_font = Font(bold=True, color="FFFFFF", size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    todos_dias = set()
    for id_emp in datos_agrupados.keys():
        todos_dias.update(datos_agrupados[id_emp]['dias'].keys())

    dias_ordenados = sorted(todos_dias)
    dias_semana = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

    dias_formateados = []
    for dia_key in dias_ordenados:
        fecha_obj = datetime.strptime(dia_key, "%Y-%m-%d")
        nombre_dia = dias_semana[fecha_obj.weekday()]
        dia_formateado = f"{nombre_dia} {fecha_obj.day}"
        dias_formateados.append((dia_key, dia_formateado))

    fila_principal = ["ID", "Nombre"]
    
    for dia_key, dia_formateado in dias_formateados:
        fila_principal.append(dia_formateado)
        fila_principal.extend(["", "", "", "", ""])

    fila_sub = ["", ""]
    for _ in dias_formateados:
        fila_sub.extend(["Entrada", "Salida", "Horas Trabajadas", "Acceso", "Foto ID", "Asistencia"])

    ws.append(fila_principal)
    ws.append(fila_sub)

    for col in range(1, 3):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    col_actual = 3
    for i in range(len(dias_formateados)):
        col_inicio = col_actual
        col_fin = col_actual + 5
        
        ws.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_fin)
        
        cell = ws.cell(1, col_inicio)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
        col_actual = col_fin + 1

    for col in range(1, len(fila_sub) + 1):
        cell = ws.cell(2, col)
        cell.fill = subheader_fill
        cell.font = subheader_font
        cell.alignment = center_align
        cell.border = thin_border

    fila_num = 3
    for id_emp in sorted(datos_agrupados.keys(), key=lambda x: int(x) if x.isdigit() else float('inf')):
        datos = datos_agrupados[id_emp]
        nombre = datos['nombre']

        fila_datos = [id_emp, nombre]

        for dia_key, dia_formateado in dias_formateados:
            if dia_key in datos['dias']:
                dia_info = datos['dias'][dia_key]
                fila_datos.append(dia_info['hora_entrada'])
                fila_datos.append(dia_info['hora_salida'] if dia_info['hora_salida'] else "")
                fila_datos.append(dia_info['horas_trabajadas'] if dia_info['horas_trabajadas'] else "")
                fila_datos.append(dia_info['tipo_acceso'])
                fila_datos.append(dia_info['foto_id'] if dia_info['foto_id'] != "No disponible" else "")
                asistencia = "Asistio" if dia_info['hora_salida'] else "Solo entrada"
                fila_datos.append(asistencia)
            else:
                fila_datos.extend(["", "", "", "", "", "No asistio"])

        ws.append(fila_datos)

        for col in range(1, len(fila_datos) + 1):
            cell = ws.cell(fila_num, col)
            cell.border = thin_border

            if col == 1:
                cell.alignment = center_align
            elif col == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

        fila_num += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20

    col = 3
    for _ in dias_formateados:
        ws.column_dimensions[get_column_letter(col)].width = 12
        ws.column_dimensions[get_column_letter(col + 1)].width = 12
        ws.column_dimensions[get_column_letter(col + 2)].width = 15
        ws.column_dimensions[get_column_letter(col + 3)].width = 10
        ws.column_dimensions[get_column_letter(col + 4)].width = 20
        ws.column_dimensions[get_column_letter(col + 5)].width = 15
        col += 6

    nombre_excel = f"Reporte_Detallado_sin_Descuentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_excel)
    logger.info(f"Reporte Detallado sin Descuentos generado: {nombre_excel}")
    return nombre_excel


def recalcular_horas_con_tolerancias(hora_entrada_real, hora_salida_real, hora_entrada_jornada, 
                                      hora_salida_jornada, dia_semana, jornada_tipo, es_servicios_generales=False, 
                                      id_empleado=None):
    """
    Recalcula las horas de entrada y salida aplicando tolerancias y reglas de horas extra.
    
    Args:
        hora_entrada_real: datetime de entrada real
        hora_salida_real: datetime de salida real (puede ser None)
        hora_entrada_jornada: str "HH:MM" de jornada
        hora_salida_jornada: str "HH:MM" de jornada
        dia_semana: int (0=Lunes, 6=Domingo)
        jornada_tipo: str "Lunes-Viernes" o "Lunes-Sabado"
        es_servicios_generales: bool, True si es empleado de servicios generales
        id_empleado: str, ID del empleado para obtener tolerancias personalizadas
    
    Returns:
        dict con entrada_ajustada, salida_ajustada, horas_extra, es_retardo
    """
    # Obtener tolerancias (personalizadas o globales)
    if id_empleado:
        tol = obtener_tolerancias_para_empleado(id_empleado)
    else:
        # Usar globales por defecto
        tol = {
            'tolerancia_entrada': TOLERANCIA_ENTRADA_TARDE,
            'tolerancia_salida_antes': TOLERANCIA_SALIDA_ANTES,
            'tolerancia_salida_despues': TOLERANCIA_SALIDA_DESPUES,
            'hora_limite_extra_lj': HORA_LIMITE_EXTRA_LJ,
            'tolerancia_hora_extra': TOLERANCIA_HORA_EXTRA,
            'horas_extra_maximo': MAX_HORAS_EXTRA,
            'viernes_permite_extra': HORAS_EXTRA_VIERNES,
            'sabado_hora_limite': SABADO_HORA_LIMITE_SALIDA,
            'sabado_tolerancia': SABADO_TOLERANCIA_SALIDA,
            'sabado_entrada_minima': SABADO_ENTRADA_MINIMA_REAL,
            'servicios_tolerancia_entrada': SERVICIOS_TOLERANCIA_ENTRADA,
            'servicios_hora_salida_min': SERVICIOS_HORA_SALIDA_MIN
        }
    
    resultado = {
        'entrada_ajustada': None,
        'salida_ajustada': None,
        'horas_extra': 0,
        'es_retardo': False,
        'observacion': ''
    }
    
    # Si no hay entrada, retornar vacío
    if not hora_entrada_real:
        return resultado
    
    # Convertir horas de jornada a datetime del mismo día
    fecha_base = hora_entrada_real.date()
    hora_jornada_entrada_dt = datetime.combine(fecha_base, 
                                               datetime.strptime(hora_entrada_jornada, "%H:%M").time())
    hora_jornada_salida_dt = datetime.combine(fecha_base, 
                                              datetime.strptime(hora_salida_jornada, "%H:%M").time())
    
    # ============= VALIDAR ENTRADA MÍNIMA (GLOBAL para todos los días) =============
    # Si entra antes de la hora mínima, siempre usar hora real (no ajustar)
    hora_entrada_minima = datetime.combine(fecha_base, 
                                          datetime.strptime(tol['sabado_entrada_minima'], "%H:%M").time())
    
    if hora_entrada_real < hora_entrada_minima:
        # Entró muy temprano → usar hora real siempre
        resultado['entrada_ajustada'] = hora_entrada_real
        resultado['observacion'] = 'Entrada temprana'
        
        # Calcular salida si existe
        if hora_salida_real:
            resultado['salida_ajustada'] = hora_salida_real
        
        return resultado
    
    # ============= REGLAS ESPECIALES SERVICIOS GENERALES =============
    if es_servicios_generales:
        # Entrada: tolerancia de 1 hora
        diferencia_entrada = (hora_entrada_real - hora_jornada_entrada_dt).total_seconds() / 60
        
        if diferencia_entrada <= 0:
            resultado['entrada_ajustada'] = hora_jornada_entrada_dt
        elif diferencia_entrada <= tol['servicios_tolerancia_entrada']:
            resultado['entrada_ajustada'] = hora_jornada_entrada_dt
        else:
            resultado['entrada_ajustada'] = hora_entrada_real
            resultado['es_retardo'] = True
        
        # Salida: puede salir desde hora configurable, siempre usar hora de jornada
        if hora_salida_real:
            hora_min_salida = datetime.combine(fecha_base, 
                                               datetime.strptime(tol['servicios_hora_salida_min'], "%H:%M").time())
            if hora_salida_real >= hora_min_salida:
                resultado['salida_ajustada'] = hora_jornada_salida_dt
                resultado['observacion'] = 'Servicios Generales'
            else:
                resultado['salida_ajustada'] = hora_salida_real
                resultado['observacion'] = 'Salida anticipada'
        
        return resultado
    
    # ============= REGLAS ESPECIALES SÁBADO =============
    if dia_semana == 5:
        # ENTRADA: Ya fue validada en sección global (entrada mínima)
        # Procesar entrada normal para sábado
        diferencia_entrada = (hora_entrada_real - hora_jornada_entrada_dt).total_seconds() / 60
        
        if diferencia_entrada <= 0:
            resultado['entrada_ajustada'] = hora_jornada_entrada_dt
        else:
            # Usar hora de jornada normal para sábado
            resultado['entrada_ajustada'] = hora_jornada_entrada_dt
        
        # SALIDA SÁBADO
        if hora_salida_real:
            # Sábado con jornada L-V = todo es hora extra
            if jornada_tipo == "Lunes-Viernes":
                # Salida máxima configurable o salida + tolerancia, lo que sea menor
                hora_limite_sabado = datetime.combine(fecha_base,
                                                     datetime.strptime(tol['sabado_hora_limite'], "%H:%M").time())
                hora_salida_mas_tolerancia = hora_salida_real + timedelta(minutes=tol['sabado_tolerancia'])
                
                # Si sale antes del límite, usar salida + tolerancia
                # Si sale después, usar máximo el límite
                if hora_salida_real <= hora_limite_sabado:
                    resultado['salida_ajustada'] = min(hora_salida_mas_tolerancia, hora_limite_sabado)
                else:
                    resultado['salida_ajustada'] = hora_limite_sabado
                
                diferencia_horas = (resultado['salida_ajustada'] - resultado['entrada_ajustada']).total_seconds() / 3600
                resultado['horas_extra'] = min(diferencia_horas, tol['horas_extra_maximo'])
                resultado['observacion'] = 'Sábado (día no laboral)'
            else:
                # Jornada L-S: aplicar reglas de salida normal pero con límite configurable
                hora_limite_sabado = datetime.combine(fecha_base,
                                                     datetime.strptime(tol['sabado_hora_limite'], "%H:%M").time())
                diferencia_salida = (hora_salida_real - hora_jornada_salida_dt).total_seconds() / 60
                
                if -tol['tolerancia_salida_antes'] <= diferencia_salida <= tol['tolerancia_salida_despues']:
                    resultado['salida_ajustada'] = hora_jornada_salida_dt
                elif hora_salida_real > hora_limite_sabado:
                    resultado['salida_ajustada'] = hora_limite_sabado
                else:
                    resultado['salida_ajustada'] = hora_salida_real
        
        return resultado
    
    # ============= CALCULAR ENTRADA (Lun-Vie) =============
    diferencia_entrada = (hora_entrada_real - hora_jornada_entrada_dt).total_seconds() / 60
    
    if diferencia_entrada <= 0:
        # Llegó antes o a tiempo → usar hora de jornada
        resultado['entrada_ajustada'] = hora_jornada_entrada_dt
    elif diferencia_entrada <= tol['tolerancia_entrada']:
        # Llegó dentro de tolerancia → usar hora de jornada
        resultado['entrada_ajustada'] = hora_jornada_entrada_dt
    elif diferencia_entrada > tol['tolerancia_entrada']:
        # Llegó tarde (más de tolerancia) → usar hora real
        resultado['entrada_ajustada'] = hora_entrada_real
        resultado['es_retardo'] = True
    
    # ============= CALCULAR SALIDA (Lun-Vie) =============
    if not hora_salida_real:
        resultado['salida_ajustada'] = None
        resultado['observacion'] = 'Solo entrada'
        return resultado
    
    diferencia_salida = (hora_salida_real - hora_jornada_salida_dt).total_seconds() / 60
    
    # VIERNES: No hay horas extra (configurable) - SIEMPRE usar hora de jornada como máximo
    if dia_semana == 4 and not tol['viernes_permite_extra']:
        if diferencia_salida <= 0:
            # Salió antes o a tiempo
            if diferencia_salida >= -tol['tolerancia_salida_antes']:
                resultado['salida_ajustada'] = hora_jornada_salida_dt
            else:
                resultado['salida_ajustada'] = hora_salida_real
                resultado['observacion'] = 'Salida anticipada'
        else:
            # Salió después de la hora de jornada - usar hora de jornada como máximo
            resultado['salida_ajustada'] = hora_jornada_salida_dt
            resultado['observacion'] = 'Viernes (sin horas extra)'
        return resultado
    
    # Salió dentro del rango permitido
    if -tol['tolerancia_salida_antes'] <= diferencia_salida <= tol['tolerancia_salida_despues']:
        resultado['salida_ajustada'] = hora_jornada_salida_dt
        resultado['horas_extra'] = 0
    
    # Salió muy temprano (más de tolerancia antes)
    elif diferencia_salida < -tol['tolerancia_salida_antes']:
        resultado['salida_ajustada'] = hora_salida_real
        resultado['observacion'] = 'Salida anticipada'
    
    # Salió tarde (más de tolerancia después)
    elif diferencia_salida > tol['tolerancia_salida_despues']:
        # LUNES-JUEVES: Calcular horas extra
        if dia_semana < 4:
            # Hora límite común configurable
            hora_limite_dt = datetime.combine(fecha_base, 
                                             datetime.strptime(tol['hora_limite_extra_lj'], "%H:%M").time())
            
            # Si salió cerca del límite (±tolerancia), ajustar al límite
            diferencia_limite = abs((hora_salida_real - hora_limite_dt).total_seconds() / 60)
            if diferencia_limite <= tol['tolerancia_hora_extra']:
                resultado['salida_ajustada'] = hora_limite_dt
            else:
                resultado['salida_ajustada'] = hora_salida_real
            
            # Calcular horas extra (máximo configurable)
            horas_trabajadas = (resultado['salida_ajustada'] - hora_jornada_salida_dt).total_seconds() / 3600
            resultado['horas_extra'] = min(horas_trabajadas, tol['horas_extra_maximo'])
        
        # SÁBADO con jornada L-S: aplicar reglas normales
        else:
            resultado['salida_ajustada'] = hora_salida_real
    
    return resultado


def procesar_asistencia_por_dia(datos_agrupados, fecha_inicio, fecha_fin):
    """
    Procesa la asistencia por día y retorna datos calculados para reutilizar en reportes
    datos_agrupados tiene estructura: {id_emp: {"nombre": ..., "dias": {fecha: {...}}}}
    Retorna: dict con estructura {fecha: {id_emp: {entrada, salida, horas_extra, estado, ...}}}
    IMPORTANTE: Muestra TODOS los empleados TODOS los días (aunque no haya registro)
    """
    from jornada_laboral import obtener_todos_empleados
    from datetime import datetime, timedelta
    
    # Obtener datos de empleados de BD (incluye jornada_tipo, etc.)
    todos_empleados = obtener_todos_empleados()
    # Datos de jornada por empleado para proyección
    empleados_info = {}
    for emp in todos_empleados:
        emp_id = emp[0]
        empleados_info[emp_id] = {
            'jornada': emp[2] if len(emp) > 2 else "",
            'lj_ent': emp[3] if len(emp) > 3 else None,
            'lj_sal': emp[4] if len(emp) > 4 else None,
            'v_ent': emp[5] if len(emp) > 5 else None,
            'v_sal': emp[6] if len(emp) > 6 else None,
            's_ent': emp[7] if len(emp) > 7 else None,
            's_sal': emp[8] if len(emp) > 8 else None,
        }
    # Mapa con jornada y horarios para proyeccion
    empleados_info = {}
    for emp in todos_empleados:
        emp_id = emp[0]
        empleados_info[emp_id] = {
            'jornada': emp[2] if len(emp) > 2 else "",
            'lj_ent': emp[3] if len(emp) > 3 else None,
            'lj_sal': emp[4] if len(emp) > 4 else None,
            'v_ent': emp[5] if len(emp) > 5 else None,
            'v_sal': emp[6] if len(emp) > 6 else None,
            's_ent': emp[7] if len(emp) > 7 else None,
            's_sal': emp[8] if len(emp) > 8 else None,
        }
    
    empleados_bd = {}
    for emp in todos_empleados:
        # Desempacar campos según la BD
        id_emp = emp[0]
        nombre = emp[1]
        jornada = emp[2]
        lj_ent = emp[3]
        lj_sal = emp[4]
        v_ent = emp[5]
        v_sal = emp[6]
        s_ent = emp[7]
        s_sal = emp[8]
        # Los campos son: 0-8 (jornadas), 9 (fecha_registro), 10 (es_servicios_generales), 11 (es_chofer)
        es_servicios_generales = emp[10] if len(emp) > 10 else 0
        es_chofer = emp[11] if len(emp) > 11 else 0
        
        empleados_bd[id_emp] = {
            'nombre': nombre,
            'jornada': jornada,
            'lj_ent': lj_ent,
            'lj_sal': lj_sal,
            'v_ent': v_ent,
            'v_sal': v_sal,
            's_ent': s_ent,
            's_sal': s_sal,
            'es_servicios_generales': bool(es_servicios_generales),
            'es_chofer': bool(es_chofer)
        }
    
    datos_procesados = {}
    
    # Generar rango de fechas (sin domingos)
    fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
    
    dias_rango = []
    fecha_actual = fecha_inicio_dt
    while fecha_actual <= fecha_fin_dt:
        # Ignorar domingos
        if fecha_actual.weekday() != 6:
            dia_key = fecha_actual.strftime("%Y-%m-%d")
            dias_rango.append((dia_key, fecha_actual))
        fecha_actual += timedelta(days=1)
    
    # Iterar sobre TODOS los empleados
    for id_emp, emp_bd in sorted(empleados_bd.items()):
        nombre = emp_bd['nombre']
        jornada = emp_bd['jornada']
        lj_ent = emp_bd['lj_ent']
        lj_sal = emp_bd['lj_sal']
        v_ent = emp_bd['v_ent']
        v_sal = emp_bd['v_sal']
        s_ent = emp_bd['s_ent']
        s_sal = emp_bd['s_sal']
        es_servicios_generales = emp_bd['es_servicios_generales']
        es_chofer = emp_bd['es_chofer']
        
        # Iterar sobre todos los días del rango
        for dia_key, fecha_obj in dias_rango:
            try:
                dia_semana = fecha_obj.weekday()  # 0=Lun, 6=Dom
                
                # Crear entrada en datos_procesados si no existe
                if dia_key not in datos_procesados:
                    datos_procesados[dia_key] = {}
                
                # Determinar hora de jornada según día de la semana
                if dia_semana < 4:  # Lunes-Jueves
                    hora_ent_jornada = lj_ent
                    hora_sal_jornada = lj_sal
                elif dia_semana == 4:  # Viernes
                    hora_ent_jornada = v_ent
                    hora_sal_jornada = v_sal
                elif dia_semana == 5:  # Sábado
                    if jornada == "Lunes-Sabado":
                        # Jornada normal de sábado
                        hora_ent_jornada = s_ent
                        hora_sal_jornada = s_sal
                    else:
                        # Lunes-Viernes pero si tiene entrada, es tiempo extra
                        if hora_entrada_real:
                            # Permitir sábado como tiempo extra: 07:00 a 17:10
                            hora_ent_jornada = "07:00"
                            hora_sal_jornada = "17:10"
                        else:
                            # No labora el sábado y no tiene entrada
                            datos_procesados[dia_key][id_emp] = {
                                'nombre': nombre,
                                'entrada': '',
                                'salida': '',
                                'horas_extra': 0,
                                'es_retardo': False,
                                'observacion': 'No labora',
                                'es_servicios_generales': es_servicios_generales,
                                'es_chofer': es_chofer
                            }
                            continue
                else:  # Domingo (ya filtrado en dias_rango)
                    continue
                
                # Obtener datos reales de asistencia del reloj (si existen)
                hora_entrada_real = None
                hora_salida_real = None
                
                if id_emp in datos_agrupados and dia_key in datos_agrupados[id_emp]['dias']:
                    dia_info = datos_agrupados[id_emp]['dias'][dia_key]
                    hora_entrada_real = dia_info.get('entrada_datetime')
                    
                    if dia_info.get('hora_salida'):
                        try:
                            hora_salida_real = datetime.strptime(
                                f"{dia_key} {dia_info['hora_salida']}",
                                "%Y-%m-%d %H:%M"
                            )
                        except:
                            pass
                
                # RECALCULAR con tolerancias (aunque no haya datos de entrada)
                recalculo = recalcular_horas_con_tolerancias(
                    hora_entrada_real, hora_salida_real,
                    hora_ent_jornada, hora_sal_jornada,
                    dia_semana, jornada, es_servicios_generales,
                    id_emp  # Pasar ID para tolerancias personalizadas
                )
                
                entrada_ajustada = ""
                salida_ajustada = ""
                
                if recalculo['entrada_ajustada']:
                    entrada_ajustada = recalculo['entrada_ajustada'].strftime("%H:%M")
                
                if recalculo['salida_ajustada']:
                    salida_ajustada = recalculo['salida_ajustada'].strftime("%H:%M")
                elif es_chofer and hora_entrada_real and not hora_salida_real:
                    # Es chofer, tiene entrada pero no salida -> "Viaje"
                    salida_ajustada = "Viaje"
                
                # Almacenar datos procesados
                datos_procesados[dia_key][id_emp] = {
                    'nombre': nombre,
                    'entrada': entrada_ajustada,
                    'salida': salida_ajustada,
                    'horas_extra': recalculo.get('horas_extra', 0),
                    'es_retardo': recalculo.get('es_retardo', False),
                    'observacion': recalculo.get('observacion', ''),
                    'es_servicios_generales': es_servicios_generales,
                    'es_chofer': es_chofer
                }
            
            except Exception as e:
                logger.error(f"Error procesando asistencia del {id_emp} en {dia_key}: {e}", exc_info=True)
                # Aun así, agregar entrada vacía para que aparezca en el reporte
                if dia_key not in datos_procesados:
                    datos_procesados[dia_key] = {}
                datos_procesados[dia_key][id_emp] = {
                    'nombre': nombre,
                    'entrada': '',
                    'salida': '',
                    'horas_extra': 0,
                    'es_retardo': False,
                    'observacion': 'Error en procesamiento',
                    'es_servicios_generales': es_servicios_generales,
                    'es_chofer': es_chofer
                }
                continue
    
    return datos_procesados


def generar_reporte_asistencia_mejorado(datos_agrupados, fecha_inicio, fecha_fin):
    """
    Genera reporte de asistencia con una hoja por día, usando datos ya procesados
    """
    try:
        from jornada_laboral import obtener_todos_empleados
    except ImportError:
        logger.warning("jornada_laboral no disponible para reporte de asistencia")
        return None

    # Procesar asistencia una sola vez
    datos_procesados = procesar_asistencia_por_dia(datos_agrupados, fecha_inicio, fecha_fin)
    
    if not datos_procesados:
        logger.warning("No hay datos procesados para generar reporte")
        return None

    wb = Workbook()
    # Eliminar la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Obtener todos los días del rango
    fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
    
    dias_del_rango = []
    fecha_actual = fecha_inicio_dt
    while fecha_actual <= fecha_fin_dt:
        # IGNORAR DOMINGOS
        if fecha_actual.weekday() != 6:
            dias_del_rango.append(fecha_actual)
        fecha_actual += timedelta(days=1)
    
    dias_semana_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    meses_nombres = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Crear una hoja por cada día
    for fecha_dia in dias_del_rango:
        dia_key = fecha_dia.strftime("%Y-%m-%d")
        dia_semana = fecha_dia.weekday()
        nombre_dia = dias_semana_nombres[dia_semana]
        nombre_mes = meses_nombres[fecha_dia.month]
        
        # Nombre de la hoja (máximo 31 caracteres para Excel)
        nombre_hoja = f"{nombre_dia[:3]} {fecha_dia.day:02d}-{fecha_dia.month:02d}"
        ws = wb.create_sheet(title=nombre_hoja)
        
        # TÍTULO: "Lunes 05 de Enero de 2025"
        titulo = f"{nombre_dia} {fecha_dia.day:02d} de {nombre_mes} de {fecha_dia.year}"
        ws.merge_cells('A1:C1')
        cell_titulo = ws['A1']
        cell_titulo.value = titulo
        cell_titulo.fill = header_fill
        cell_titulo.font = header_font
        cell_titulo.alignment = center_align
        cell_titulo.border = thin_border
        
        # ENCABEZADOS: Nombre | Entrada | Salida | Permiso Entrada | Permiso Salida
        ws['A2'] = 'Nombre'
        ws['B2'] = 'Entrada'
        ws['C2'] = 'Salida'
        ws['D2'] = 'Permiso Entrada'
        ws['E2'] = 'Permiso Salida'
        
        for col in ['A2', 'B2', 'C2', 'D2', 'E2']:
            cell = ws[col]
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = thin_border
        
        # DATOS de asistencia usando datos ya procesados
        fila = 3
        
        # Si no hay datos para este día, saltarlo
        if dia_key not in datos_procesados:
            continue
        
        # Obtener permisos del día
        from jornada_laboral import obtener_permisos_dia
        permisos_dia = {}
        try:
            permisos = obtener_permisos_dia(dia_key)
            for perm in permisos:
                id_emp_perm = perm[1]  # id_empleado está en índice 1
                tipo_permiso = perm[3]  # tipo_permiso está en índice 3
                hora_permiso = perm[4]  # hora está en índice 4
                if id_emp_perm not in permisos_dia:
                    permisos_dia[id_emp_perm] = {}
                permisos_dia[id_emp_perm][tipo_permiso] = hora_permiso
        except Exception as e:
            logger.warning(f"Error obteniendo permisos para {dia_key}: {e}")
        
        for id_emp in sorted(datos_procesados[dia_key].keys()):
            emp_data = datos_procesados[dia_key][id_emp]
            nombre = emp_data['nombre']
            entrada_str = emp_data['entrada']
            salida_str = emp_data['salida']
            
            # Obtener permisos para este empleado si existen
            permiso_entrada = permisos_dia.get(id_emp, {}).get('entrada', '')
            permiso_salida = permisos_dia.get(id_emp, {}).get('salida', '')
            
            # Escribir fila
            ws[f'A{fila}'] = nombre
            ws[f'B{fila}'] = entrada_str
            ws[f'C{fila}'] = salida_str
            ws[f'D{fila}'] = permiso_entrada
            ws[f'E{fila}'] = permiso_salida
            
            # Aplicar estilos
            ws[f'A{fila}'].alignment = left_align
            ws[f'B{fila}'].alignment = center_align
            ws[f'C{fila}'].alignment = center_align
            ws[f'D{fila}'].alignment = center_align
            ws[f'E{fila}'].alignment = center_align
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{fila}'].border = thin_border
            
            fila += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    # Guardar archivo
    nombre_archivo = f"Reporte_Asistencia_Detallado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_archivo)
    logger.info(f"Reporte de asistencia detallado generado: {nombre_archivo}")
    return nombre_archivo


def generar_reporte_asistencia(datos_agrupados):
    """
    Genera reporte de asistencia considerando jornada laboral
    """
    try:
        from jornada_laboral import obtener_todos_empleados
    except ImportError:
        logger.warning("jornada_laboral no disponible, generando reporte sin validacion de jornada")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    asistio_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    tarde_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    falta_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    solo_entrada_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    todos_empleados = obtener_todos_empleados()
    
    todos_dias = set()
    for id_emp in datos_agrupados.keys():
        todos_dias.update(datos_agrupados[id_emp]['dias'].keys())

    # Filtrar domingos
    dias_ordenados = []
    for dia_key in sorted(todos_dias):
        fecha_obj = datetime.strptime(dia_key, "%Y-%m-%d")
        if fecha_obj.weekday() != 6:  # No incluir domingos
            dias_ordenados.append(dia_key)
    
    dias_semana_nombres = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]

    fila_encabezado = ["ID", "Nombre"]
    for dia_key in dias_ordenados:
        fecha_obj = datetime.strptime(dia_key, "%Y-%m-%d")
        nombre_dia = dias_semana_nombres[fecha_obj.weekday()]
        fila_encabezado.append(f"{nombre_dia} {fecha_obj.day}")

    ws.append(fila_encabezado)

    for col in range(1, len(fila_encabezado) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    fila_num = 2
    for emp_row in todos_empleados:
        id_emp = emp_row[0]
        nombre = emp_row[1]
        jornada = emp_row[2]
        lj_ent = emp_row[3]
        lj_sal = emp_row[4]
        v_ent = emp_row[5]
        v_sal = emp_row[6]
        s_ent = emp_row[7]
        s_sal = emp_row[8]
        
        fila_datos = [id_emp, nombre]
        
        for dia_key in dias_ordenados:
            fecha_obj = datetime.strptime(dia_key, "%Y-%m-%d")
            dia_semana = fecha_obj.weekday()

            if dia_semana < 4:
                hora_entrada_esp = lj_ent
                hora_salida_esp = lj_sal
            elif dia_semana == 4:
                hora_entrada_esp = v_ent
                hora_salida_esp = v_sal
            elif dia_semana == 5:
                if jornada == "Lunes-Sabado":
                    hora_entrada_esp = s_ent
                    hora_salida_esp = s_sal
                else:
                    fila_datos.append("No labora")
                    continue
            else:
                fila_datos.append("No labora")
                continue

            estado = "Falta"
            color = falta_fill

            if id_emp in datos_agrupados and dia_key in datos_agrupados[id_emp]['dias']:
                dia_info = datos_agrupados[id_emp]['dias'][dia_key]
                hora_entrada_real = dia_info['hora_entrada']
                hora_salida_real = dia_info['hora_salida']

                if hora_salida_real:
                    estado = "Asistio"
                    color = asistio_fill
                    
                    if hora_entrada_real > hora_entrada_esp:
                        estado = "Tarde"
                        color = tarde_fill
                else:
                    estado = "Solo entrada"
                    color = solo_entrada_fill

            fila_datos.append(estado)

        ws.append(fila_datos)

        for col_idx, valor in enumerate(fila_datos, 1):
            cell = ws.cell(fila_num, col_idx)
            cell.border = thin_border

            if col_idx == 1:
                cell.alignment = center_align
            elif col_idx == 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

                if "Asistio" in str(valor):
                    cell.fill = asistio_fill
                elif "Tarde" in str(valor):
                    cell.fill = tarde_fill
                elif "Solo entrada" in str(valor):
                    cell.fill = solo_entrada_fill
                elif "Falta" in str(valor):
                    cell.fill = falta_fill

        fila_num += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    for col in range(3, len(fila_encabezado) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    nombre_archivo = f"Reporte_Asistencia_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_archivo)
    logger.info(f"Reporte de asistencia generado: {nombre_archivo}")
    return nombre_archivo


def calcular_horas_jornada(hora_inicio_str, hora_fin_str, descontar_desayuno=True, id_empleado=None):
    """Devuelve horas planeadas descontando desayuno y comida.
    
    Args:
        hora_inicio_str: Hora de inicio en formato "HH:MM"
        hora_fin_str: Hora de fin en formato "HH:MM"
        descontar_desayuno: Si False, no descuenta los minutos de desayuno (para administrativos)
        id_empleado: ID del empleado para usar tolerancias personalizadas (opcional)
    
    Returns:
        float: Horas calculadas
    """
    try:
        if not hora_inicio_str or not hora_fin_str:
            return 0
        base_date = datetime(2000, 1, 1)
        inicio = datetime.strptime(hora_inicio_str, "%H:%M")
        fin = datetime.strptime(hora_fin_str, "%H:%M")
        inicio_dt = base_date.replace(hour=inicio.hour, minute=inicio.minute)
        fin_dt = base_date.replace(hour=fin.hour, minute=fin.minute)
        horas = (fin_dt - inicio_dt).total_seconds() / 3600

        # Desayuno: solo descontar si está dentro de la jornada Y descontar_desayuno es True
        if descontar_desayuno:
            desayuno_inicio = base_date.replace(hour=int(DESAYUNO_HORA.split(":")[0]), minute=int(DESAYUNO_HORA.split(":")[1]))
            desayuno_fin = desayuno_inicio + timedelta(minutes=DESAYUNO_MINUTOS_DESCUENTO)
            # Solo descontar si el desayuno cae dentro del rango de trabajo
            if inicio_dt < desayuno_fin and fin_dt > desayuno_inicio:
                horas -= DESAYUNO_MINUTOS_DESCUENTO / 60

        # Comida condicional si hay traslape
        # IMPORTANTE: Solo descontar comida si la salida es DESPUÉS de (comida_inicio + tolerancia)
        comida_inicio = base_date.replace(hour=int(COMIDA_HORA_INICIO.split(":")[0]), minute=0)
        
        # Obtener tolerancia personalizada si existe
        comida_tol = COMIDA_TOLERANCIA_SALIDA
        if id_empleado:
            tol = obtener_tolerancias_para_empleado(id_empleado)
            comida_tol = tol.get('comida_tolerancia_salida', COMIDA_TOLERANCIA_SALIDA)
        
        comida_inicio_con_tolerancia = comida_inicio + timedelta(minutes=comida_tol)
        comida_fin = comida_inicio + timedelta(minutes=COMIDA_DURACION_MINUTOS)
        # La comida solo se descuenta si sale DESPUÉS de (comida_inicio + tolerancia)
        # Ej: si comida a 13:00 y tolerancia 10 min → descuenta si sale después de 13:10
        if fin_dt > comida_inicio_con_tolerancia and inicio_dt < comida_fin:
            horas -= COMIDA_DURACION_MINUTOS / 60

        return max(0, horas)
    except Exception:
        return 0


def generar_reporte_nomina(datos_agrupados, fecha_inicio, fecha_fin):
    """
    Genera reporte de nómina con hoja principal de horas trabajadas y hojas de detalle
    Primera hoja: Resumen con Nombre, Horas Trabajadas por día, Horas Totales
    Hojas adicionales: Asistencia detallada igual que reporte asistencia
    """
    logger.info("Generando reporte de nómina...")
    
    from jornada_laboral import obtener_todos_empleados
    
    wb = Workbook()
    
    # ===== ESTILOS =====
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=9)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    number_align = Alignment(horizontal="right", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    todos_empleados = obtener_todos_empleados()
    # Datos de jornada por empleado para proyección
    empleados_info = {}
    for emp in todos_empleados:
        emp_id = emp[0]
        empleados_info[emp_id] = {
            'jornada': emp[2] if len(emp) > 2 else "",
            'lj_ent': emp[3] if len(emp) > 3 else None,
            'lj_sal': emp[4] if len(emp) > 4 else None,
            'v_ent': emp[5] if len(emp) > 5 else None,
            'v_sal': emp[6] if len(emp) > 6 else None,
            's_ent': emp[7] if len(emp) > 7 else None,
            's_sal': emp[8] if len(emp) > 8 else None,
        }

    # Generar lista de fechas (excluyendo domingos)
    fecha_init_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin_obj = datetime.strptime(fecha_fin, "%Y-%m-%d")
    
    fechas_lista = []
    fecha_actual = fecha_init_obj
    first_friday = None
    first_saturday = None
    while fecha_actual <= fecha_fin_obj:
        # Excluir domingos (weekday 6)
        if fecha_actual.weekday() != 6:
            fecha_str = fecha_actual.strftime("%Y-%m-%d")
            fechas_lista.append(fecha_str)
            if fecha_actual.weekday() == 4 and first_friday is None:
                first_friday = fecha_str
            if fecha_actual.weekday() == 5 and first_saturday is None:
                first_saturday = fecha_str
        fecha_actual += timedelta(days=1)
    # Últimos viernes y sábados del rango para proyección
    last_friday = None
    last_saturday = None
    for f in fechas_lista:
        dow = datetime.strptime(f, "%Y-%m-%d").weekday()
        if dow == 4:
            last_friday = f
        elif dow == 5:
            last_saturday = f
    dow_abbr = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    
    # ===== HOJA CONFIGURACION (para formulas) =====
    ws_config = wb.create_sheet("Config Nómina")
    ws_config["A1"] = "Desayuno Hora"
    ws_config["B1"] = DESAYUNO_HORA
    ws_config["A2"] = "Desayuno Minutos"
    ws_config["B2"] = DESAYUNO_MINUTOS_DESCUENTO
    ws_config["A3"] = "Comida Hora Inicio"
    # Guardar solo la hora como entero para TIME()
    try:
        ws_config["B3"] = int(COMIDA_HORA_INICIO.split(":")[0])
    except Exception:
        ws_config["B3"] = 13
    ws_config["A4"] = "Comida Duración Minutos"
    ws_config["B4"] = COMIDA_DURACION_MINUTOS
    ws_config.column_dimensions['A'].width = 28
    ws_config.column_dimensions['B'].width = 10

    # ===== HOJA 1: RESUMEN DE HORAS =====
    ws_resumen = wb.active
    ws_resumen.title = "Horas Trabajadas"
    
    # Encabezado principal (fila 1)
    ws_resumen['A1'] = "Nombre"
    ws_resumen['A1'].fill = header_fill
    ws_resumen['A1'].font = header_font
    ws_resumen['A1'].alignment = center_align
    ws_resumen['A1'].border = thin_border
    ws_resumen.column_dimensions['A'].width = 25
    
    # Rango de columnas para días
    first_day_col_idx = 2
    last_day_col_idx = first_day_col_idx + len(fechas_lista) - 1 if fechas_lista else first_day_col_idx
    if fechas_lista:
        ws_resumen.merge_cells(start_row=1, start_column=first_day_col_idx, end_row=1, end_column=last_day_col_idx)
        ws_resumen.cell(row=1, column=first_day_col_idx, value="HORAS TRABAJADAS").fill = header_fill
        ws_resumen.cell(row=1, column=first_day_col_idx).font = header_font
        ws_resumen.cell(row=1, column=first_day_col_idx).alignment = center_align
        ws_resumen.cell(row=1, column=first_day_col_idx).border = thin_border
    
    # Columna de Total (al final)
    col_total_idx = last_day_col_idx + 1
    col_total = get_column_letter(col_total_idx)
    ws_resumen[f'{col_total}1'] = "TOTAL HORAS"
    ws_resumen[f'{col_total}1'].fill = header_fill
    ws_resumen[f'{col_total}1'].font = header_font
    ws_resumen[f'{col_total}1'].alignment = center_align
    ws_resumen[f'{col_total}1'].border = thin_border
    ws_resumen.column_dimensions[col_total].width = 14

    # Columna de Proyección
    col_proj_idx = col_total_idx + 1
    col_proj = get_column_letter(col_proj_idx)
    ws_resumen[f'{col_proj}1'] = "HORAS PROYECCION"
    ws_resumen[f'{col_proj}1'].fill = header_fill
    ws_resumen[f'{col_proj}1'].font = header_font
    ws_resumen[f'{col_proj}1'].alignment = center_align
    ws_resumen[f'{col_proj}1'].border = thin_border
    ws_resumen.column_dimensions[col_proj].width = 16

    # Columna de Notas
    col_notes_idx = col_proj_idx + 1
    col_notes = get_column_letter(col_notes_idx)
    ws_resumen[f'{col_notes}1'] = "NOTAS"
    ws_resumen[f'{col_notes}1'].fill = header_fill
    ws_resumen[f'{col_notes}1'].font = header_font
    ws_resumen[f'{col_notes}1'].alignment = center_align
    ws_resumen[f'{col_notes}1'].border = thin_border
    ws_resumen.column_dimensions[col_notes].width = 30
    
    # Subencabezados de fechas (fila 2)
    col_idx = first_day_col_idx
    dias_semana_abbr_header = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    for fecha in fechas_lista:
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
        nombre_dia = dias_semana_abbr_header[fecha_obj.weekday()]
        fecha_formato = f"{nombre_dia} {fecha_obj.day}/{fecha_obj.month}"
        
        col_letter = get_column_letter(col_idx)
        ws_resumen[f'{col_letter}2'] = fecha_formato
        ws_resumen[f'{col_letter}2'].fill = subheader_fill
        ws_resumen[f'{col_letter}2'].font = subheader_font
        ws_resumen[f'{col_letter}2'].alignment = center_align
        ws_resumen[f'{col_letter}2'].border = thin_border
        ws_resumen.column_dimensions[col_letter].width = 12
        
        col_idx += 1
    
    # Llenar datos de empleados (desde fila 3)
    fila = 3
    for emp_index, emp in enumerate(sorted(todos_empleados, key=lambda e: e[0])):
        emp_id = emp[0]
        emp_nombre = emp[1]
        emp_info = empleados_info.get(emp_id, {})
        jornada_tipo = emp_info.get('jornada', '')
        
        notes_parts = []
        ws_resumen[f'A{fila}'] = emp_nombre
        ws_resumen[f'A{fila}'].font = Font(bold=True)
        ws_resumen[f'A{fila}'].alignment = left_align
        ws_resumen[f'A{fila}'].border = thin_border
        
        # Horas por cada día
        col_idx = first_day_col_idx
        skip_cells = []
        
        for fecha in fechas_lista:
            # Nombre de hoja igual al reporte de asistencia: "Lun 05-12"
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
            dias_semana_nombres = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
            nombre_dia_abbr = dias_semana_nombres[fecha_obj.weekday()]
            sheet_name = f"{nombre_dia_abbr} {fecha_obj.day:02d}-{fecha_obj.month:02d}"
            
            # Fila en hoja del día: encabezados en 2, datos desde 3
            row_in_day_sheet = 3 + emp_index
            entrada_ref = f"'{sheet_name}'!B{row_in_day_sheet}"
            salida_ref = f"'{sheet_name}'!C{row_in_day_sheet}"
            desayuno_min_ref = "'Config Nómina'!B2"
            comida_hora_ref = "'Config Nómina'!B3"
            comida_min_ref = "'Config Nómina'!B4"

            # Identificar si es el primer viernes/sábado a excluir del total (pero mantener visible)
            is_skip = False
            if jornada_tipo == "Lunes-Viernes" and first_friday and fecha == first_friday:
                is_skip = True
            if jornada_tipo == "Lunes-Sabado" and first_friday and fecha == first_friday:
                is_skip = True
            if jornada_tipo == "Lunes-Sabado" and first_saturday and fecha == first_saturday:
                is_skip = True

            formula = (
                f"=IF(AND({entrada_ref}<>\"\", {salida_ref}=\"\"), \"Asistió\", "
                f"IF({entrada_ref}=\"\", 0, "
                f"(((({salida_ref}-{entrada_ref})*24) - {desayuno_min_ref}/60 - "
                f"IF(AND({entrada_ref} < TIME({comida_hora_ref},0,0) + {comida_min_ref}/1440, "
                f"{salida_ref} > TIME({comida_hora_ref},0,0)), {comida_min_ref}/60, 0))/24)))"
            )

            col_letter = get_column_letter(col_idx)
            ws_resumen[f'{col_letter}{fila}'] = formula
            ws_resumen[f'{col_letter}{fila}'].alignment = number_align
            ws_resumen[f'{col_letter}{fila}'].border = thin_border
            ws_resumen[f'{col_letter}{fila}'].number_format = '[h]:mm'

            if is_skip:
                skip_cells.append(f"{col_letter}{fila}")
                # Si es el primer viernes (y hubo salida) comparar contra jornada para posible nota de descuento
                if fecha == first_friday and jornada_tipo in ("Lunes-Viernes", "Lunes-Sabado"):
                    horas_jornada_vie = calcular_horas_jornada(emp_info.get('v_ent'), emp_info.get('v_sal'))
                    if horas_jornada_vie > 0:
                        # convertir referencia a decimal horas para comparar
                        current_cell = f"{col_letter}{fila}"
                        # Nota: en Excel, *24 convierte la fracción de día a horas
                        notes_parts.append(
                            f"IF(AND({entrada_ref}<>\"\", {salida_ref}<>\"\", ({current_cell}*24)<{horas_jornada_vie}),"
                            f"\"Descuento viernes -\"&TEXT(( {horas_jornada_vie}-({current_cell}*24) )/24,\"[h]:mm\"),\"\")"
                        )

            # El total será formula de SUM al final; aquí no sumamos valor fijo
            col_idx += 1
        
        # Total horas
        first_col_letter = get_column_letter(first_day_col_idx)
        last_col_letter = get_column_letter(last_day_col_idx)
        base_sum = f"SUM({first_col_letter}{fila}:{last_col_letter}{fila})"
        if skip_cells:
            skip_sum = "+".join(skip_cells)
            total_expr = f"={base_sum}-SUM({skip_sum})"
        else:
            total_expr = f"={base_sum}"

        ws_resumen[f'{col_total}{fila}'] = total_expr
        ws_resumen[f'{col_total}{fila}'].font = Font(bold=True)
        ws_resumen[f'{col_total}{fila}'].alignment = number_align
        ws_resumen[f'{col_total}{fila}'].border = thin_border
        ws_resumen[f'{col_total}{fila}'].number_format = '[h]:mm'

        # Proyección: sumar jornada completa de viernes (y sábado si aplica) si no hay salida registrada; descontar sábado ausente
        emp_info = empleados_info.get(emp_id, {})
        jornada_tipo = emp_info.get('jornada', '')

        proy_parts = []
        # Friday projection
        if last_friday:
            friday_row = 3 + emp_index
            fri_dt = datetime.strptime(last_friday, "%Y-%m-%d")
            friday_sheet = f"{dow_abbr[fri_dt.weekday()]} {fri_dt.day:02d}-{fri_dt.month:02d}"
            friday_entry = f"'{friday_sheet}'!B{friday_row}"
            friday_exit = f"'{friday_sheet}'!C{friday_row}"
            horas_viernes = calcular_horas_jornada(emp_info.get('v_ent'), emp_info.get('v_sal'))
            if horas_viernes > 0:
                # Si hay entrada y no hay salida, proyecta jornada completa; si hay entrada y salida, no suma nada; si no hay entrada, no suma
                proy_parts.append(
                    f"IF({friday_entry}<>\"\", IF({friday_exit}=\"\", {horas_viernes}/24, 0), 0)"
                )
                notes_parts.append(
                    f"IF({friday_entry}<>\"\", IF({friday_exit}=\"\", \"Proy viernes +\"&TEXT({horas_viernes}/24,\"[h]:mm\"), \"\"), \"\")"
                )
        # Saturday projection for L-S
        if jornada_tipo == "Lunes-Sabado" and last_saturday:
            saturday_row = 3 + emp_index
            sat_dt = datetime.strptime(last_saturday, "%Y-%m-%d")
            saturday_sheet = f"{dow_abbr[sat_dt.weekday()]} {sat_dt.day:02d}-{sat_dt.month:02d}"
            saturday_entry = f"'{saturday_sheet}'!B{saturday_row}"
            saturday_exit = f"'{saturday_sheet}'!C{saturday_row}"
            horas_sabado = calcular_horas_jornada(emp_info.get('s_ent'), emp_info.get('s_sal'))
            if horas_sabado > 0:
                # Si hay entrada y no salida, proyecta jornada completa; si hay entrada y salida, no suma; si no hay entrada, descuenta jornada
                proy_parts.append(
                    f"IF({saturday_entry}<>\"\", IF({saturday_exit}=\"\", {horas_sabado}/24, 0), -{horas_sabado}/24)"
                )
                notes_parts.append(
                    f"IF({saturday_entry}<>\"\", IF({saturday_exit}=\"\", \"Proy sábado +\"&TEXT({horas_sabado}/24,\"[h]:mm\"), \"\"), \"Sábado ausente -\"&TEXT({horas_sabado}/24,\"[h]:mm\"))"
                )

        sum_range = base_sum
        if skip_cells:
            sum_range = f"{sum_range}-SUM({'+'.join(skip_cells)})"
        if proy_parts:
            proy_formula = "=" + sum_range + "+" + "+".join(proy_parts)
        else:
            proy_formula = "=" + sum_range

        ws_resumen[f'{col_proj}{fila}'] = proy_formula
        ws_resumen[f'{col_proj}{fila}'].font = Font(bold=True)
        ws_resumen[f'{col_proj}{fila}'].alignment = number_align
        ws_resumen[f'{col_proj}{fila}'].border = thin_border
        ws_resumen[f'{col_proj}{fila}'].number_format = '[h]:mm'

        # Notas
        if notes_parts:
            ws_resumen[f'{col_notes}{fila}'] = "=" + "&".join([f"{n}" for n in notes_parts])
        ws_resumen[f'{col_notes}{fila}'].alignment = left_align
        ws_resumen[f'{col_notes}{fila}'].border = thin_border
        
        fila += 1
    
    # ===== HOJA 2: RESUMEN DE HORAS EN FORMATO DECIMAL =====
    ws_decimal = wb.create_sheet("Horas Trabajadas (Decimal)")
    
    # Encabezado principal (fila 1)
    ws_decimal['A1'] = "Nombre"
    ws_decimal['A1'].fill = header_fill
    ws_decimal['A1'].font = header_font
    ws_decimal['A1'].alignment = center_align
    ws_decimal['A1'].border = thin_border
    ws_decimal.column_dimensions['A'].width = 25
    
    # Rango de columnas para días (igual que la hoja de hora)
    col_idx = first_day_col_idx
    if fechas_lista:
        ws_decimal.merge_cells(start_row=1, start_column=first_day_col_idx, end_row=1, end_column=last_day_col_idx)
        ws_decimal.cell(row=1, column=first_day_col_idx, value="HORAS TRABAJADAS (Decimal)").fill = header_fill
        ws_decimal.cell(row=1, column=first_day_col_idx).font = header_font
        ws_decimal.cell(row=1, column=first_day_col_idx).alignment = center_align
        ws_decimal.cell(row=1, column=first_day_col_idx).border = thin_border
    
    # Columna de Total
    ws_decimal[f'{col_total}1'] = "TOTAL HORAS"
    ws_decimal[f'{col_total}1'].fill = header_fill
    ws_decimal[f'{col_total}1'].font = header_font
    ws_decimal[f'{col_total}1'].alignment = center_align
    ws_decimal[f'{col_total}1'].border = thin_border
    ws_decimal.column_dimensions[col_total].width = 14

    # Columna de Proyección
    ws_decimal[f'{col_proj}1'] = "HORAS PROYECCION"
    ws_decimal[f'{col_proj}1'].fill = header_fill
    ws_decimal[f'{col_proj}1'].font = header_font
    ws_decimal[f'{col_proj}1'].alignment = center_align
    ws_decimal[f'{col_proj}1'].border = thin_border
    ws_decimal.column_dimensions[col_proj].width = 16

    # Columna de Notas
    ws_decimal[f'{col_notes}1'] = "NOTAS"
    ws_decimal[f'{col_notes}1'].fill = header_fill
    ws_decimal[f'{col_notes}1'].font = header_font
    ws_decimal[f'{col_notes}1'].alignment = center_align
    ws_decimal[f'{col_notes}1'].border = thin_border
    ws_decimal.column_dimensions[col_notes].width = 30
    
    # Subencabezados de fechas (fila 2)
    col_idx = first_day_col_idx
    for fecha in fechas_lista:
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
        nombre_dia = dias_semana_abbr_header[fecha_obj.weekday()]
        fecha_formato = f"{nombre_dia} {fecha_obj.day}/{fecha_obj.month}"
        
        col_letter = get_column_letter(col_idx)
        ws_decimal[f'{col_letter}2'] = fecha_formato
        ws_decimal[f'{col_letter}2'].fill = subheader_fill
        ws_decimal[f'{col_letter}2'].font = subheader_font
        ws_decimal[f'{col_letter}2'].alignment = center_align
        ws_decimal[f'{col_letter}2'].border = thin_border
        ws_decimal.column_dimensions[col_letter].width = 12
        
        col_idx += 1
    
    # Llenar datos de empleados en formato decimal (desde fila 3)
    fila = 3
    for emp_index, emp in enumerate(sorted(todos_empleados, key=lambda e: e[0])):
        emp_id = emp[0]
        emp_nombre = emp[1]
        emp_info = empleados_info.get(emp_id, {})
        jornada_tipo = emp_info.get('jornada', '')
        
        ws_decimal[f'A{fila}'] = emp_nombre
        ws_decimal[f'A{fila}'].font = Font(bold=True)
        ws_decimal[f'A{fila}'].alignment = left_align
        ws_decimal[f'A{fila}'].border = thin_border
        
        # Horas por cada día (en formato decimal, referenciando la hoja de hora)
        col_idx = first_day_col_idx
        skip_cells_decimal = []
        
        for fecha in fechas_lista:
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
            nombre_dia_abbr = dias_semana_nombres[fecha_obj.weekday()]
            sheet_name = f"{nombre_dia_abbr} {fecha_obj.day:02d}-{fecha_obj.month:02d}"
            row_in_day_sheet = 3 + emp_index
            
            # Referenciar celda de la hoja de hora y convertir a decimal (* 24 porque está en formato de fracción de día)
            col_letter = get_column_letter(col_idx)
            cell_ref = f"'{sheet_name}'!{col_letter}{fila}"
            
            # Identificar si es el primer viernes/sábado a excluir del total
            is_skip = False
            if jornada_tipo == "Lunes-Viernes" and first_friday and fecha == first_friday:
                is_skip = True
            if jornada_tipo == "Lunes-Sabado" and first_friday and fecha == first_friday:
                is_skip = True
            if jornada_tipo == "Lunes-Sabado" and first_saturday and fecha == first_saturday:
                is_skip = True
            
            # Copiar valor de la hoja hora, convertir a decimal
            ws_decimal[f'{col_letter}{fila}'] = f"=IF('{ws_resumen.title}'!{col_letter}{fila}=\"Asistió\", \"Asistió\", IF(ISNUMBER('{ws_resumen.title}'!{col_letter}{fila}), '{ws_resumen.title}'!{col_letter}{fila}*24, \"{ws_resumen.title}'!{col_letter}{fila}\"))"
            ws_decimal[f'{col_letter}{fila}'].alignment = number_align
            ws_decimal[f'{col_letter}{fila}'].border = thin_border
            ws_decimal[f'{col_letter}{fila}'].number_format = '0.00'  # Formato decimal con 2 decimales
            
            if is_skip:
                skip_cells_decimal.append(f"{col_letter}{fila}")
            
            col_idx += 1
        
        # Total horas en formato decimal
        first_col_letter = get_column_letter(first_day_col_idx)
        last_col_letter = get_column_letter(last_day_col_idx)
        base_sum_decimal = f"SUM({first_col_letter}{fila}:{last_col_letter}{fila})"
        if skip_cells_decimal:
            skip_sum_decimal = "+".join(skip_cells_decimal)
            total_expr_decimal = f"={base_sum_decimal}-SUM({skip_sum_decimal})"
        else:
            total_expr_decimal = f"={base_sum_decimal}"

        ws_decimal[f'{col_total}{fila}'] = total_expr_decimal
        ws_decimal[f'{col_total}{fila}'].font = Font(bold=True)
        ws_decimal[f'{col_total}{fila}'].alignment = number_align
        ws_decimal[f'{col_total}{fila}'].border = thin_border
        ws_decimal[f'{col_total}{fila}'].number_format = '0.00'
        
        # Proyección en decimal
        proy_parts_decimal = []
        if last_friday:
            friday_row = 3 + emp_index
            fri_dt = datetime.strptime(last_friday, "%Y-%m-%d")
            friday_sheet = f"{dow_abbr[fri_dt.weekday()]} {fri_dt.day:02d}-{fri_dt.month:02d}"
            friday_entry = f"'{friday_sheet}'!B{friday_row}"
            friday_exit = f"'{friday_sheet}'!C{friday_row}"
            horas_viernes = calcular_horas_jornada(emp_info.get('v_ent'), emp_info.get('v_sal'))
            if horas_viernes > 0:
                proy_parts_decimal.append(
                    f"IF({friday_entry}<>\"\", IF({friday_exit}=\"\", {horas_viernes}, 0), 0)"
                )
        if jornada_tipo == "Lunes-Sabado" and last_saturday:
            saturday_row = 3 + emp_index
            sat_dt = datetime.strptime(last_saturday, "%Y-%m-%d")
            saturday_sheet = f"{dow_abbr[sat_dt.weekday()]} {sat_dt.day:02d}-{sat_dt.month:02d}"
            saturday_entry = f"'{saturday_sheet}'!B{saturday_row}"
            saturday_exit = f"'{saturday_sheet}'!C{saturday_row}"
            horas_sabado = calcular_horas_jornada(emp_info.get('s_ent'), emp_info.get('s_sal'))
            if horas_sabado > 0:
                proy_parts_decimal.append(
                    f"IF({saturday_entry}<>\"\", IF({saturday_exit}=\"\", {horas_sabado}, 0), -{horas_sabado})"
                )
        
        sum_range_decimal = base_sum_decimal
        if skip_cells_decimal:
            sum_range_decimal = f"{sum_range_decimal}-SUM({'+'.join(skip_cells_decimal)})"
        if proy_parts_decimal:
            proy_formula_decimal = "=" + sum_range_decimal + "+" + "+".join(proy_parts_decimal)
        else:
            proy_formula_decimal = "=" + sum_range_decimal

        ws_decimal[f'{col_proj}{fila}'] = proy_formula_decimal
        ws_decimal[f'{col_proj}{fila}'].font = Font(bold=True)
        ws_decimal[f'{col_proj}{fila}'].alignment = number_align
        ws_decimal[f'{col_proj}{fila}'].border = thin_border
        ws_decimal[f'{col_proj}{fila}'].number_format = '0.00'
        
        # Copiar notas de la hoja principal
        ws_decimal[f'{col_notes}{fila}'] = f"='{ws_resumen.title}'!{col_notes}{fila}"
        ws_decimal[f'{col_notes}{fila}'].alignment = left_align
        ws_decimal[f'{col_notes}{fila}'].border = thin_border
        
        fila += 1
    
    # ===== HOJAS DE ASISTENCIA POR DIA =====
    generar_hojas_asistencia(wb, datos_agrupados, fechas_lista)
    
    nombre_archivo = f"Reporte_Nomina_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(nombre_archivo)
    logger.info(f"Reporte de nómina generado: {nombre_archivo}")
    return nombre_archivo


def calcular_horas_empleado_dia(emp_id, fecha, datos_agrupados):
    """
    Calcula las horas trabajadas de un empleado en un día específico.
    Descuenta desayuno y comida (considerando tolerancias personalizadas).
    """
    if emp_id not in datos_agrupados:
        return 0
    
    if fecha not in datos_agrupados[emp_id]['dias']:
        return 0
    
    dia_data = datos_agrupados[emp_id]['dias'][fecha]
    
    # Si no hay salida, no hay horas trabajadas
    if not dia_data.get('hora_entrada_real') or not dia_data.get('hora_salida_real'):
        return 0
    
    try:
        entrada = datetime.strptime(dia_data['hora_entrada_real'], "%H:%M")
        salida = datetime.strptime(dia_data['hora_salida_real'], "%H:%M")
        
        # Calcular horas totales
        tiempo_trabajado = (salida - entrada).total_seconds() / 3600  # en horas
        
        # Descontar desayuno (siempre 10 minutos)
        tiempo_trabajado -= DESAYUNO_MINUTOS_DESCUENTO / 60
        
        # Descontar comida (si entra antes y sale después de la hora de comida + tolerancia)
        comida_inicio = datetime.strptime(COMIDA_HORA_INICIO, "%H:%M")
        
        # Obtener tolerancia personalizada si existe
        comida_tol = COMIDA_TOLERANCIA_SALIDA
        tol = obtener_tolerancias_para_empleado(emp_id)
        comida_tol = tol.get('comida_tolerancia_salida', COMIDA_TOLERANCIA_SALIDA)
        
        comida_inicio_con_tolerancia = comida_inicio + timedelta(minutes=comida_tol)
        comida_duracion_horas = COMIDA_DURACION_MINUTOS / 60
        comida_fin = comida_inicio.replace(hour=comida_inicio.hour, minute=comida_inicio.minute) + timedelta(minutes=COMIDA_DURACION_MINUTOS)
        
        # Si el empleado trabajó durante la hora de comida (considerando tolerancia de salida), descontar
        # Ej: si comida a 13:00 y tolerancia 10 min → descuenta si sale después de 13:10
        if entrada < comida_fin and salida > comida_inicio_con_tolerancia:
            tiempo_trabajado -= comida_duracion_horas
        
        return max(0, tiempo_trabajado)  # No permitir horas negativas
    except:
        return 0


def generar_hojas_asistencia(wb, datos_agrupados, fechas_lista):
    """
    Genera hojas por día exactamente como en el reporte de asistencia:
    Título en A1:C1 con "<Día> <dd> de <Mes> de <YYYY>" y encabezados en A2:C2
    Columnas: Nombre | Entrada | Salida
    """
    from jornada_laboral import obtener_todos_empleados
    
    # Procesar asistencia una vez para obtener entradas/salidas consolidadas
    fecha_inicio = fechas_lista[0] if fechas_lista else datetime.now().strftime("%Y-%m-%d")
    fecha_fin = fechas_lista[-1] if fechas_lista else fecha_inicio
    datos_procesados = procesar_asistencia_por_dia(datos_agrupados, fecha_inicio, fecha_fin)

    # Estilos similares al mejorado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if not datos_procesados:
        return

    # Diccionarios de nombres de día/mes en español
    dias_semana_nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    dias_semana_abbr = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    meses_nombres = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                     "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    for fecha in fechas_lista:
        fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
        dia_key = fecha
        dia_semana = fecha_obj.weekday()
        nombre_dia = dias_semana_nombres[dia_semana]
        nombre_dia_abbr = dias_semana_abbr[dia_semana]
        nombre_mes = meses_nombres[fecha_obj.month]

        # Nombre hoja como en asistencia mejorado
        nombre_hoja = f"{nombre_dia_abbr} {fecha_obj.day:02d}-{fecha_obj.month:02d}"
        ws = wb.create_sheet(title=nombre_hoja)

        # Título
        titulo = f"{nombre_dia} {fecha_obj.day:02d} de {nombre_mes} de {fecha_obj.year}"
        ws.merge_cells('A1:C1')
        cell_titulo = ws['A1']
        cell_titulo.value = titulo
        cell_titulo.fill = header_fill
        cell_titulo.font = header_font
        cell_titulo.alignment = center_align
        cell_titulo.border = thin_border

        # Encabezados
        ws['A2'] = 'Nombre'
        ws['B2'] = 'Entrada'
        ws['C2'] = 'Salida'
        ws['D2'] = 'Permiso Entrada'
        ws['E2'] = 'Permiso Salida'
        for col in ['A2', 'B2', 'C2', 'D2', 'E2']:
            cell = ws[col]
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = center_align
            cell.border = thin_border

        fila = 3
        # Si no hay datos para el día, continuar con hoja vacía
        if dia_key not in datos_procesados:
            # Ajustar ancho incluso en hojas vacías
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            continue

        # Obtener permisos del día
        from jornada_laboral import obtener_permisos_dia
        permisos_dia = {}
        try:
            permisos = obtener_permisos_dia(dia_key)
            for perm in permisos:
                id_emp_perm = perm[1]  # id_empleado está en índice 1
                tipo_permiso = perm[3]  # tipo_permiso está en índice 3
                hora_permiso = perm[4]  # hora está en índice 4
                if id_emp_perm not in permisos_dia:
                    permisos_dia[id_emp_perm] = {}
                permisos_dia[id_emp_perm][tipo_permiso] = hora_permiso
        except Exception as e:
            logger.warning(f"Error obteniendo permisos para {dia_key}: {e}")

        # Orden por ID estable
        for id_emp in sorted(datos_procesados[dia_key].keys()):
            emp_data = datos_procesados[dia_key][id_emp]
            nombre = emp_data['nombre']
            entrada_str = emp_data['entrada']
            salida_str = emp_data['salida']
            
            # Obtener permisos para este empleado si existen
            permiso_entrada = permisos_dia.get(id_emp, {}).get('entrada', '')
            permiso_salida = permisos_dia.get(id_emp, {}).get('salida', '')

            # Escribir nombre
            ws[f'A{fila}'] = nombre
            ws[f'A{fila}'].alignment = left_align
            ws[f'A{fila}'].border = thin_border
            ws[f'B{fila}'] = entrada_str
            ws[f'B{fila}'].alignment = center_align
            ws[f'B{fila}'].border = thin_border
            ws[f'C{fila}'] = salida_str
            ws[f'C{fila}'].alignment = center_align
            ws[f'C{fila}'].border = thin_border
            ws[f'D{fila}'] = permiso_entrada
            ws[f'D{fila}'].alignment = center_align
            ws[f'D{fila}'].border = thin_border
            ws[f'E{fila}'] = permiso_salida
            ws[f'E{fila}'].alignment = center_align
            ws[f'E{fila}'].border = thin_border

            # Convertir a tiempo Excel si posible
            def to_time(val):
                if not val:
                    return None
                for fmt in ("%H:%M", "%H:%M:%S"):
                    try:
                        return datetime.strptime(val, fmt).time()
                    except Exception:
                        continue
                return None

            entrada_time = to_time(entrada_str)
            salida_time = to_time(salida_str)

            if entrada_time:
                ws[f'B{fila}'] = entrada_time
                ws[f'B{fila}'].number_format = "hh:mm"
            else:
                ws[f'B{fila}'] = ""
            ws[f'B{fila}'].alignment = center_align
            ws[f'B{fila}'].border = thin_border

            if salida_time:
                ws[f'C{fila}'] = salida_time
                ws[f'C{fila}'].number_format = "hh:mm"
            else:
                ws[f'C{fila}'] = ""
            ws[f'C{fila}'].alignment = center_align
            ws[f'C{fila}'].border = thin_border

            fila += 1

        # Ajuste de anchos como en asistencia
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    # No retornar nada


def generar_reporte(fecha_inicio=None, fecha_fin=None):
    logger.info("=" * 60)
    logger.info("INICIANDO GENERACION DE REPORTE DE ASISTENCIA")
    logger.info("=" * 60)

    # Si no se proporcionan fechas, usar fechas por defecto
    if not fecha_inicio:
        fecha_inicio = "2025-12-01"
    if not fecha_fin:
        fecha_fin = datetime.now().strftime("%Y-%m-%d")
    
    # Formato para el reloj: YYYY-MM-DDTHH:MM:SS
    fecha_inicio_reloj = f"{fecha_inicio}T00:00:00"
    
    # SIEMPRE agregar 1 día a la fecha final para consultar al reloj
    # Esto asegura capturar todos los eventos del día seleccionado
    fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")
    fecha_fin_reloj_dt = fecha_fin_dt + timedelta(days=1)
    fecha_fin_reloj = fecha_fin_reloj_dt.strftime("%Y-%m-%dT23:59:59")
    
    logger.info(f"Rango de fechas seleccionado: {fecha_inicio} a {fecha_fin}")
    logger.info(f"Consultando al reloj hasta: {fecha_fin_reloj} (día siguiente para capturar eventos completos)")

    try:
        auth = HTTPDigestAuth(USER, PASS)

        logger.info("Paso 1: Sincronizando nombres de empleados...")
        nombres_map = obtener_nombres_empleados(auth)

        if not nombres_map:
            logger.warning("No se cargo la base de datos de empleados. Continuando...")

        time.sleep(1)

        logger.info("Paso 2: Descargando eventos del reloj...")
        lista_asistencia = []
        fotos_pendientes = []
        posicion = 0
        bloque = 35
        search_id = "fetch_" + datetime.now().strftime("%H%M%S")

        while True:
            payload = {
                "AcsEventCond": {
                    "searchID": search_id,
                    "searchResultPosition": posicion,
                    "maxResults": bloque,
                    "major": 0,
                    "minor": 0,
                    "startTime": fecha_inicio_reloj,
                    "endTime": fecha_fin_reloj
                }
            }
            try:
                res = requests.post(
                    f"https://{IP_RELOJ}/ISAPI/AccessControl/AcsEvent?format=json",
                    json=payload, auth=auth, verify=False, timeout=15
                )

                if res.status_code != 200:
                    logger.error(f"Error al descargar eventos: Status {res.status_code}")
                    break

                data = res.json().get('AcsEvent', {})
                eventos = data.get('InfoList', [])

                if not eventos:
                    logger.info("No hay mas eventos disponibles")
                    break

                for ev in eventos:
                    id_emp = str(ev.get('employeeNoString') or ev.get('employeeNo') or '')
                    if not id_emp or id_emp == '0':
                        continue

                    m_code = ev.get('minor', 0)

                    if m_code == 75:
                        tipo_acceso = "ROSTRO"
                    elif m_code == 38:
                        tipo_acceso = "HUELLA"
                    elif m_code == 1:
                        tipo_acceso = "TARJETA"
                    else:
                        logger.debug(f"Evento filtrado - ID: {id_emp}, Minor code: {m_code}")
                        continue

                    fecha_raw = ev.get('time', '')
                    logger.debug(f"Evento capturado - ID: {id_emp}, Fecha: {fecha_raw}, Tipo: {tipo_acceso}")
                    url_pic = ev.get('pictureURL')

                    f_safe = fecha_raw.replace("-", "").replace(":", "").replace("T", "_").split("+")[0].replace("Z", "")
                    nombre_jpg = f"{id_emp}_{f_safe}.jpg"

                    status_foto = "No disponible"
                    if url_pic:
                        fotos_pendientes.append({
                            'url': url_pic,
                            'archivo': nombre_jpg,
                            'id_emp': id_emp
                        })
                        status_foto = nombre_jpg

                    lista_asistencia.append({
                        "Fecha y Hora": fecha_raw,
                        "ID": id_emp,
                        "Nombre": nombres_map.get(id_emp, f"ID: {id_emp}"),
                        "Tipo de Acceso": tipo_acceso,
                        "Archivo Foto": status_foto
                    })

                num_matches = data.get('numOfMatches', 0)
                if num_matches == 0 or (posicion + num_matches) >= data.get('totalMatches', 0):
                    break

                posicion += num_matches
                logger.info(f"Progreso: {len(lista_asistencia)} registros descargados...")

            except requests.exceptions.Timeout:
                logger.error("Timeout al descargar eventos")
                break
            except requests.exceptions.ConnectionError as e:
                logger.error(f"Error de conexion descargando eventos: {e}")
                break
            except ValueError as e:
                logger.error(f"Error al parsear JSON de eventos: {e}")
                break
            except Exception as e:
                logger.error(f"Error inesperado descargando eventos: {type(e).__name__} - {e}")
                break

        logger.info(f"Total de registros descargados: {len(lista_asistencia)}")
        
        # FILTRAR eventos para mostrar solo hasta la fecha final seleccionada (no incluir el día extra)
        if lista_asistencia:
            fecha_fin_limite = datetime.strptime(fecha_fin, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            lista_asistencia_filtrada = []
            
            for reg in lista_asistencia:
                try:
                    fecha_evento = pd.to_datetime(reg['Fecha y Hora']).tz_localize(None)
                    if fecha_evento <= fecha_fin_limite:
                        lista_asistencia_filtrada.append(reg)
                except:
                    # Si hay error al parsear la fecha, incluir el registro
                    lista_asistencia_filtrada.append(reg)
            
            eventos_filtrados = len(lista_asistencia) - len(lista_asistencia_filtrada)
            if eventos_filtrados > 0:
                logger.info(f"Filtrados {eventos_filtrados} eventos posteriores a {fecha_fin} 23:59:59")
            
            lista_asistencia = lista_asistencia_filtrada
            logger.info(f"Registros a procesar después del filtro: {len(lista_asistencia)}")
        
        # Mostrar resumen de últimos eventos capturados
        if lista_asistencia:
            logger.info("=" * 60)
            logger.info("ÚLTIMOS 5 EVENTOS CAPTURADOS:")
            for reg in lista_asistencia[-5:]:
                logger.info(f"  ID: {reg['ID']:8} | {reg['Nombre']:20} | {reg['Fecha y Hora']:19} | {reg['Tipo de Acceso']}")
            logger.info("=" * 60)

        if fotos_pendientes:
            logger.info(f"Paso 3: Descargando {len(fotos_pendientes)} fotos en paralelo...")
            descargar_fotos_paralelo(fotos_pendientes, auth)
        else:
            logger.info("Paso 3: No hay fotos para descargar")

        if lista_asistencia:
            logger.info("Paso 4: Procesando entrada y salida...")
            datos_procesados = procesar_entrada_salida(lista_asistencia)

            logger.info("Paso 5: Generando Excel detallado...")
            nombre_excel = generar_excel_final(datos_procesados)
            
            logger.info("Paso 6: Generando reporte de asistencia mejorado...")
            nombre_reporte_asistencia = generar_reporte_asistencia_mejorado(datos_procesados, fecha_inicio, fecha_fin)
            
            logger.info("Paso 7: Generando reporte de nómina...")
            nombre_reporte_nomina = generar_reporte_nomina(datos_procesados, fecha_inicio, fecha_fin)
            
            logger.info("=" * 60)
            logger.info("EXITO! Reportes generados correctamente")
            logger.info(f"Reporte detallado: {nombre_excel}")
            if nombre_reporte_asistencia:
                logger.info(f"Reporte asistencia: {nombre_reporte_asistencia}")
            if nombre_reporte_nomina:
                logger.info(f"Reporte nomina: {nombre_reporte_nomina}")
            logger.info(f"Fotos guardadas en: {CARPETA_FOTOS}")
            logger.info("=" * 60)
        else:
            logger.warning("No hay registros para generar el reporte")

    except Exception as e:
        logger.critical(f"Error critico en la generacion del reporte: {type(e).__name__} - {e}")
        raise


if __name__ == "__main__":
    generar_reporte()